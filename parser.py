import xml.etree.ElementTree as ET
from utils import *
from openpyxl import load_workbook

#********************************************************* Application functions ********************************************************#
def handleXlsx(xlsxFile):
    #Load excel file with openpyxl load_workbook
    workbook = load_workbook(filename=xlsxFile)
    sheet = workbook.active

    #Convert excel sheet into a list of dictionary with header:value pairs
    reader = ExcelDictReader(sheet)

    #Convert dictionary list into conversionMap
    conversionMap = {}
    for row in reader:

        funcParams = row['function_parameters'].split('\n')
        newfuncParamsPair = []

        for param in funcParams:
            param = param.split('=')
            paramNameText = {
                'name': param[0],
                'text': param[1]
            }
            newfuncParamsPair.append(paramNameText)

        conversionMap[row['teststep.desc old']] = {
            'description': row['test_step.desc new'],
            'function_library': row['function_library'],
            'function_name': row['function_name'],
            'function_parameters': newfuncParamsPair,
        }

    return conversionMap

    
def getTestStepData(xmlInFile, conversionMap):
    tree = ET.parse(xmlInFile)
    root = tree.getroot()
    childParentMap = {child: parent for parent in root.iter() for child in parent}
    allTestSteps = root.iter('teststep')
    
    # Initialize xmlData list
    xmlData = []
    
    for index, teststep in enumerate(allTestSteps):
        
        # Get teststep description attribute from teststep
        oldDesciption = teststep.get('desc')
        
        if oldDesciption in conversionMap:
            oldFunctionLibrary = teststep.find('function_library').text
            oldFunctionName = teststep.find('function_name').text
            oldFunctionParams = teststep.find('function_parameters').iter('param')
            
            # Create old data object
            oldTestStepData = {
                'description': oldDesciption,
                'function_library': oldFunctionLibrary,
                'function_name': oldFunctionName,
                'function_parameters': [{
                    'name': param.get('name'),
                    'text': param.text.strip('\n ')
                    } for param in oldFunctionParams]
            }
            
            # Create new data object
            newTestStepData = {
                'description': conversionMap[oldDesciption]['description'],
                'function_library': conversionMap[oldDesciption]['function_library'],
                'function_name': conversionMap[oldDesciption]['function_name'],
                'function_parameters': [{
                    'name': param['name'].strip(),
                    'text': param['text'].strip('\n ')
                    } for param in conversionMap[oldDesciption]['function_parameters']]
            }
            
            # Append to teststep data list
            xmlData.append({
                'id': index + 1,
                'parentId': childParentMap[teststep].get('id'),
                'parentName': childParentMap[teststep].get('name'),
                'old': oldTestStepData,
                'new': newTestStepData,
            })
        
    return xmlData


def convertXML(filteredIds, xmlInFile, xmlOutFile, conversionMap):
    tree = ET.parse(xmlInFile)
    root = tree.getroot()
    counter = 0
    
    # Create a teststep object for every teststep in the xml 
    allTestSteps = [{'id': index + 1, 'teststep': teststep} 
                    for index, teststep in enumerate(root.iter('teststep'))]

    for teststep in allTestSteps:
       
        # Get teststep description attribute from teststep
        oldtestStepDescription = teststep['teststep'].get('desc')

        # If teststep description finds match in conversionMap - convert to new version
        if oldtestStepDescription in conversionMap and teststep['id'] in filteredIds:
            counter += 1
            matchedConfig = conversionMap[oldtestStepDescription]

            #Change old teststep description to new description
            teststep['teststep'].set('desc', matchedConfig['description'])

            #Change old function_library text to new function_library text
            oldFunctionLibrary = teststep['teststep'].find('function_library')
            oldFunctionLibrary.text = matchedConfig['function_library']

            #Change old function_name text to new function_name text
            oldFunctionName = teststep['teststep'].find('function_name')
            oldFunctionName.text = matchedConfig['function_name']

            #Delete old function parameters and replace with new ones from conversionMap
            oldFunctionParams = teststep['teststep'].find('function_parameters')

            #remove existing param elements from function_parameters
            for param in list(oldFunctionParams.iter('param')):
                oldFunctionParams.remove(param)
               
            #create new param elements and append to function_parameters
            for param in matchedConfig['function_parameters']:
                newParam = ET.SubElement(oldFunctionParams, 'param')
                newParam.set('name', param['name'])
                newParam.text = param['text']

            # Debug print
            print(f"Converted teststeps: {counter}_______________________________________________________________________________________________________________________________")
            print(f'''
id: {teststep['id']}
{teststep['teststep'].get('desc')}
{oldFunctionLibrary.text}
{oldFunctionName.text}
{[f"{param.get('name')}={param.text}" for param in oldFunctionParams]}
            ''')
    
    # Write modified xml file to specificed file location
    tree.write(xmlOutFile)



#********************************************************* Test functions ********************************************************#
def testHandleXlsx():
    xlsxFile = './testdata/mapping.xlsx'
    
    conversionMap = handleXlsx(xlsxFile)
    for key, value in conversionMap.items():
        print(key)
        for key, value in value.items():
            print(f'{key}: {value}')
            print()
            
        break


def testHandleConvertXML():
    xlsxFile = './testdata/config.xlsx'
    xmlFile = './testdata/input.xml'

    conversionMap = handleXlsx(xlsxFile)
    convertXML(xmlFile, './testdata/output.xml', conversionMap)
    
    
def testHandleGetTestStepData():
    xlsxFile = './testdata/config.xlsx'
    xmlFile = './testdata/input.xml'
    
    conversionMap = handleXlsx(xlsxFile)
    
    for item in getTestStepData(xmlFile, conversionMap):
        for key, value in item.items():
            print(f'{key}: {value}')
            print()
        print('___________________________________________________________________________________________')
    
    
    
if __name__ == '__main__':
    testHandleGetTestStepData()
