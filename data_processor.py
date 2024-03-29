from typing import Set
import xml.etree.ElementTree as ET
import re
import collections
import openpyxl
from utils import removeWhiteSpace, ExcelDictReader
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.utils import quote_sheetname
import os
import config_warnings

baseDir = os.path.dirname(__file__)

# * Define DD2.0 config excel columns
HEADER_COLUMN_MAP = {
    'description': 'C',
    'function_library': 'D',
    'function_name': 'E',
    'function_parameters': 'F'
}

# ********************************************************* Application functions ********************************************************#

''' 
Handles processing of the ATP conversion mapping by reading the mapping sheet of the excel config.

The function receives the config excel file path string set by the user through the GUI,
referenceMap object returned by handleReferenceData method,
functionDefinitionMap object returned by handleFunctionDefintionData method.

Returns a conversionMap object containing the teststep translation,
Returns a keywordMap object containing the keyword mapping of the teststeps
Returns a list of warning objects raised by the config excel.
'''
def handleMappingData(xlsxFile, referenceMap, functionDefinitionMap):
    # Load excel file with openpyxl load_workbook
    workbook = load_workbook(filename=xlsxFile)
    sheet = workbook['mapping']

    # Convert excel sheet into a list of dictionary with header:value pairs
    reader = ExcelDictReader(sheet)

    conversionMap = {}
    keywordMap = {}
    duplicateDescriptionkeys = {
        'title': 'Duplicate description keys',
        'warning': config_warnings.DUP_DESCRIPTION_KEYS_WARNING,
        'worksheet': 'mapping',
        'data': []
    }
    duplicateKeywords = {
        'title': 'Duplicate keyword set',
        'warning': config_warnings.DUP_KEYWORD_SETS_WARNING,
        'worksheet': 'mapping',
        'data': []
    }
    invalidReferenceKeys = {
        'title': 'Invalid reference key',
        'warning': config_warnings.INVALID_REFERENCE_KEYS_WARNING,
        'worksheet': 'mapping',
        'data': []
    }
    invalidTranslations = {
        'title': 'Invalid DD2.0 tranlsations',
        'warning': config_warnings.INVALID_DD2_TRANSLATIONS,
        'worksheet': 'mapping',
        'data': []
    }

    # * Create mapping based on each row
    for rowCount, row in enumerate(reader):

        # Get keywords from each row
        keywords = tuple(set([
            removeWhiteSpace(keyword.lower())
            for keyword in row["keywords"].split("\n")
            if keyword
        ]))

        # Get classic description from each row
        oldDescriptions = [
            description
            for description in row["classic teststep description"].split("\n")
            if description
        ]

        # Convert classic description to lower case to ensure case insensitive matching
        cleanedOldDescriptions = [
            removeWhiteSpace(description.lower()) for description in oldDescriptions
        ]

        # * Get DD2.0 translations
        new_description = row["DD2 teststep description"].strip()
        new_function_library = row["DD2 function_library"].strip()
        new_function_name = row["DD2 function_name"].strip()
        new_function_parameters = []

        # Create a list of name to text function parameters
        funcParams = [
            param.strip()
            for param in row["DD2 function_parameters"].split("\n")
            if param
        ]

        # * Iterate through list of function parameters and split name to text by '='
        for param in funcParams:
            param_data = [item.strip() for item in param.split("=")]

            # * If name and text is unpacked, create function parameter obj and append to new function parameter list
            if len(param_data) == 2:
                new_function_parameters.append({
                    "name": removeWhiteSpace(param_data[0]),
                    "text": removeWhiteSpace(param_data[1])
                })
            # * If only 1 value is unpacked, check if value is a reference which startswith and endswith '##'
            elif re.match(r'^#{2}.+#{2}$', param_data[0]):
                referenceKey = removeWhiteSpace(param_data[0].strip('#').lower())
                referenceData = referenceMap.get(referenceKey)

                if referenceData:
                    for data in referenceData:
                        new_function_parameters.append(data)
                else:
                    invalidReferenceKeys['data'].append({
                        f"##{referenceKey}##": 'F' + str(rowCount + 2)
                    })

        # * Generate conversionMap and duplicate description key data
        # * If classic description field is empty, append a empty_description_key
        if not cleanedOldDescriptions:
            cleanedOldDescriptions.append(f"empty_description_key - [row: {rowCount+2}]")
            oldDescriptions.append(f"empty_description_key - [row: {rowCount+2}]")

        # * Iterate through cleanded and original description together, add first instance of mapping to conversionMap
        # * else append it to the duplicate description keys list
        for cleanedOldDescription, oldDescription in zip(cleanedOldDescriptions, oldDescriptions):

            # * Check if classic description key has not been matched if true, add to conversionMap
            if cleanedOldDescription not in conversionMap:
                conversionMap[cleanedOldDescription] = {
                    "isMatched": True,  # matching flag to check if mapping has been matched in the xml input
                    "configRowCount": rowCount + 2,  # Store excel row number of mapping
                    "oldDescription": oldDescription,
                    "description": new_description,
                    "function_library": new_function_library,
                    "function_name": new_function_name,
                    "function_parameters": new_function_parameters,
                }
            # * Else add duplicate classic key to duplicateDescriptionKeys for alert
            else:
                duplicateDescriptionkeys['data'].append({
                    oldDescription: 'A' + str(rowCount + 2)
                })

        # * If there are keywords specified 
        # * Generate keyword map and duplicate key word data
        if keywords:
            # * Add first instance of keyword tuple to keywordMap
            # * else if keyword tuple already exist, append it to duplicate keyword list
            if keywords not in keywordMap:
                keywordMap[keywords] = {
                    "isMatched": False,  # matching flag to check if mapping has been matched in the xml input
                    "configRowCount": rowCount + 2,  # Store excel row number of mapping
                    "oldDescription": oldDescription,
                    "description": new_description,
                    "function_library": new_function_library,
                    "function_name": new_function_name,
                    "function_parameters": new_function_parameters,
                }
            else:
                duplicateKeywords['data'].append({
                    str(keywords): 'A' + str(rowCount + 2)
                })

        # * If there are any invalid DD2.0 function translation used
        # * Add it to invalidTranslations data
        if functionDefinitionMap:
            
            if not new_function_library:
                continue
            function_library_data = functionDefinitionMap.get(new_function_library)
            if function_library_data is None:
                invalidTranslations['data'].append({
                    new_function_library: 'D' + str(rowCount + 2)
                })
                continue
            
            if not new_function_name:
                continue
            function_name_data = function_library_data.get(new_function_name)
            if function_name_data is None:
                invalidTranslations['data'].append({
                    new_function_name: 'E' + str(rowCount + 2)
                })
                continue

            function_parameter_data = function_name_data['function_parameters']
            for param in new_function_parameters:
                if param['name'] not in function_parameter_data:
                    invalidTranslations['data'].append({
                        param['name']: 'F' + str(rowCount + 2)
                    })

    emptyFields = getEmptyFieldData(conversionMap)

    return (
        conversionMap, 
        keywordMap, 
        [
            emptyFields,
            duplicateDescriptionkeys,
            duplicateKeywords, 
            invalidReferenceKeys,
            invalidTranslations
        ]
    )

'''
Handles the processing of the reference mapping by reading the parameter reference sheet of the excel config.

The function receives the config excel file path string set by the user through the GUI.

Returns a referenceMap object to be used by the handleMappingData function,
Returns a duplicateReferences object containing warnings raised by duplicate reference keys.
'''
def handleReferenceData(xlsxFile):
    # Load excel file with openpyxl load_workbook
    workbook = load_workbook(filename=xlsxFile)
    sheet = workbook['parameter references']

    # Convert excel sheet into a list of dictionary with header:value pairs
    reader = ExcelDictReader(sheet)

    # generate reference map by iterating through each row of key and values
    referenceMap = {}
    duplicateReferences = {
        'title': 'Duplicate reference keys',
        'warning': config_warnings.DUP_REFERENCE_KEYS_WARNINGS,
        'worksheet': 'parameter references',
        'data': []
    }

    for rowCount, row in enumerate(reader):
        key = removeWhiteSpace(row['key'].lower().strip('#'))
        if not key: continue

        values = [item.strip() for item in row['values'].split('\n') if item]

        processedValues = []
        for value in values:
            value = [item.strip() for item in value.split("=")]
            if len(value) == 2:
                processedValues.append({
                    "name": value[0],
                    "text": value[1]
                })

        if key not in referenceMap:
            referenceMap[key] = processedValues
        else:
            duplicateReferences['data'].append({
                f"##{key}##": 'A' + str(rowCount + 2)
            })

    return referenceMap, duplicateReferences

'''
Handles the processing of the function definition database by reading the __ATPFunctionDefinitions.xlsx.

The function receives the hardcoded path string to the function definition database.

Returns a functionDefinitionMap object to be used by the handleMappingData function,
Returns a duplicateFunctionNames object containing warnings raised by duplicate functions.
'''
def handleFunctionDefinitionData(xlsxFile):
    # Load excel file with openpyxl load_workbook
    workbook = load_workbook(filename=xlsxFile)
    sheet = workbook['function definitions']

    # Convert excel sheet into a list of dictionary with header:value pairs
    reader = ExcelDictReader(sheet)
    functionDefinitionMap = collections.defaultdict(lambda: {})
    duplicateFunctionNames = {
        'title': 'Duplicate function name',
        'warning': config_warnings.DUP_FUNCTION_NAMES_WARNINGS,
        'worksheet': 'function definitions',
        'data': {}
    }
    functionNames = set()

    for rowCount, row in enumerate(reader):
        function_library = row['function_library']
        function_name = row['function_name']
        function_parameters = [name.strip() for name in row['function_parameters'].split('\n') if name]

        if function_library and function_name not in functionNames :
            functionDefinitionMap[function_library][function_name] = {
                'rowCount': rowCount + 2,
                'function_parameters': function_parameters
            }
            functionNames.add(function_name)
        elif function_library and function_name:
            duplicateFunctionNames['data'][function_name] = 'B' + str(rowCount + 2)

    return functionDefinitionMap, duplicateFunctionNames

'''
Handles the parsing and translation of the classical ATP teststeps to the DD2.0 capable version.

The function recieves the ATP XML file path set by the user through the GUI,
conversionMap object returned by the handleMappingData function,
keywordMap object returned by the handleMappinData function.

Returns a testcaseSortedXmlData object which contains the teststep translation data sorted by its testcase
Returns a updated conversionMap object which includes information on which teststeps were matched in the provided xml file
'''
def handleTranslateXmlData(xmlInFile, conversionMap, keywordMap):
    tree = ET.parse(xmlInFile)
    root = tree.getroot()
    childParentMap = {child: parent for parent in root.iter()
                      for child in parent}  
    allTestSteps = root.iter("teststep")

    # Initialize xmlData list
    xmlData = []

    for index, teststep in enumerate(allTestSteps):

        # Get teststep description attribute from teststep
        oldDescription = teststep.get("desc")
        cleanedOldDescription = removeWhiteSpace(oldDescription.lower())

        # * If teststep description matches a mapping in conversionMap, generate data object and append to xmlData
        if cleanedOldDescription in conversionMap:
            # If teststep description finds match in conversionMap - set isMatched to True
            mapping = conversionMap[cleanedOldDescription]
            mapping["isMatched"] = True

            xmlData.append(
                getTeststepData(
                    index, teststep, mapping, 
                    cleanedOldDescription, oldDescription, 
                    childParentMap
                )
            )

        # * If a teststep description matches a particular set of keywords, generate data object and append to xmlData
        for keys, mapping in keywordMap.items():

            # if a key does not match the description skip generate data object for the mapping
            isAllMatched = True
            for key in keys:
                if key not in cleanedOldDescription:
                    isAllMatched = False

            # if there keywords does not match, skip generation of teststep data
            if not isAllMatched: 
                continue

            # or if teststep data is already generate, skip generation of teststep data
            if next((item for item in xmlData if item['id'] == index+1), None) is not None:
                continue

            xmlData.append(
                getTeststepData(
                    index, teststep, mapping, 
                    cleanedOldDescription, oldDescription, 
                    childParentMap
                )
            )

    # * Filter teststeps into their respective testcases
    testcaseSortedXmlData = collections.defaultdict(lambda: [])
    if xmlData:
        # Filter each teststep into their respect testcases
        for teststep in xmlData:
            testcaseSortedXmlData[teststep['parentId']].append(teststep)

    return testcaseSortedXmlData, conversionMap

'''
Handles the parsing of unmatched classic teststep description keys found in the config excel file.

The function recevies a convsersionMap object returned by the handleTranslateXmlData function which includes information
on which classical teststep description key was matched

Returns a list of unmatched teststep description keys
'''
def getUnmatchedClassicDescriptions(conversionMap):
    unmatchedClassicDescriptions = []

    for index, (key, mapping) in enumerate(conversionMap.items()):
        if mapping['isMatched'] == False and not key.startswith('empty_description_key'):
            unmatchedClassicDescriptions.append(f"{mapping['oldDescription']} - [row: {index+2}]")
    
    return unmatchedClassicDescriptions

'''
Handles the conversion of the selected classical teststeps to their DD2.0 saves a converted copy of the ATP xml file at a user specifed file path.

The function receives a set of filteredIds to indicate which teststeps are selected for conversion,
xmlinFile string specifying the file path of the input xml file
xmlOutFile string specifying the file path of the converted xml file
conversionMap object which specifies the updated DD2.0 translation mappings for the teststeps

No return value, writes new xml file to xmlOutfile path
'''
def handleConvertXml(filteredIds, xmlInFile, xmlOutFile, conversionMap):
    tree = ET.parse(xmlInFile)
    root = tree.getroot()
    counter = 0

    # Create a teststep object for every teststep in the xml
    allTestSteps = [
        {"id": index + 1, "teststep": teststep}
        for index, teststep in enumerate(root.iter("teststep"))
    ]

    print("Displaying converted teststeps")
    for index, teststep in enumerate(allTestSteps):

        # Get teststep description attribute from teststep
        oldtestStepDescription = teststep["teststep"].get("desc")
        oldtestStepDescription = f"ID: {index+1} - {oldtestStepDescription}"

        # If teststep description finds match in conversionMap - convert to new version
        if oldtestStepDescription in conversionMap and teststep["id"] in filteredIds:
            counter += 1
            matchedConfig = conversionMap[oldtestStepDescription]

            # Change old teststep description to new description
            teststep["teststep"].set("desc", matchedConfig["description"])

            # Change old function_library text to new function_library text
            oldFunctionLibrary = teststep["teststep"].find("function_library")
            oldFunctionLibrary.text = matchedConfig["function_library"]

            # Change old function_name text to new function_name text
            oldFunctionName = teststep["teststep"].find("function_name")
            oldFunctionName.text = matchedConfig["function_name"]

            # Delete old function parameters and replace with new ones from conversionMap
            oldFunctionParams = teststep["teststep"].find(
                "function_parameters")

            # remove existing param elements from function_parameters
            for param in list(oldFunctionParams.iter("param")):
                oldFunctionParams.remove(param)

            # create new param elements and append to function_parameters
            for name, text in matchedConfig["function_parameters"].items():
                newParam = ET.SubElement(oldFunctionParams, "param")
                newParam.set("name", name)
                newParam.text = text

            # Debug print
            print(
                f"""
id: {teststep['id']}
{teststep['teststep'].get('desc')}
{oldFunctionLibrary.text}
{oldFunctionName.text}
{[f"{param.get('name')}={param.text}" for param in oldFunctionParams]}
            """
            )

            print(
                f"Converted teststeps: {counter} ______________________________________________________________________________________________"
            )

    # Write modified xml file to specificed file location
    tree.write(xmlOutFile)

'''
Handles the updating of the config excel file to the mapping sheet with the changes made in the GUI,
this includes the updating of DD2.0 translations and the data validation for the function information columns.

The function receives the updated functionDefinitionMap that is configured through the GUI,
xlsxInfile string specifying the config excel file path to be updated,
xlsxOutFile string specifying the config excel file path to be written to,
configData object which contains the latest config mapping updated through the GUI.

No return value, writes new config excel file to xlsxOutFile
'''
def handleConfigFileUpdate(functionDefinitionMap, xlsxInFile, xlsxOutFile, configData):

    # * Load excel file with openpyxl load_workbook
    workbook = load_workbook(filename=xlsxInFile)
    mappingSheet = workbook['mapping']

    # * If config data is generated by application
    # * Update config excel with the new mapping data generated from the application UI
    if configData:
        for configRowCount, mapping in configData.items():

            # * Update config excel mapping translation with new data
            for header, column in HEADER_COLUMN_MAP.items():
                cell = mappingSheet[column + str(configRowCount)]

                if header == 'function_parameters':
                    cell.value = '\n'.join(mapping[header])
                else:
                    cell.value = mapping[header]

    # * Clear all exisitng data validations in mapping sheet
    mappingSheet.data_validations = DataValidationList()
    
    # * Data validation for DD2.0 function library
    library_dv = DataValidation(
        type='list', 
        formula1=f"{quote_sheetname('function definitions')}!$E$2:$E${len(functionDefinitionMap)+1 if functionDefinitionMap else '1048576'}",
        allow_blank=False,
        showInputMessage=False,
        showErrorMessage=True,
        errorStyle='warning',
        errorTitle='Invalid function library',
        error='Function library entered is not specifed in ATP function definitions.'
    )
    mappingSheet.add_data_validation(library_dv)
    library_dv.add(f"D2:D1048576")

    # * Data validation for DD2.0 function name
    name_dv = DataValidation(
        type='list',
        formula1=f"=INDIRECT(D2)", 
        allow_blank=False,
        showInputMessage=False,
        showErrorMessage=True,
        errorStyle='warning',
        errorTitle='Invalid function name',
        error='Function name entered is not defined under selected function library.'
    )
    mappingSheet.add_data_validation(name_dv)
    name_dv.add(f"E2:E1048576")

    # * Data validation for DD2.0 function parameters
    param_dv = DataValidation(
        type='list', 
        formula1=f"=INDIRECT(E2)", 
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=False,
        promptTitle='Function parameters',
        prompt='name1=value\nname2=value\nname3=value1,value2'
    )
    mappingSheet.add_data_validation(param_dv)
    param_dv.add(f"F2:F1048576")

    # * Get function definitions sheet from workbook
    try:
        functionLibrarySheet = workbook['function definitions']
    except KeyError:
        # * If workbook does not contain function definitions sheet
        # * Create function definitions sheet
        functionLibrarySheet = workbook.create_sheet('function definitions')
        functionLibrarySheet.sheet_state = 'hidden'
        functionLibrarySheet.protection.sheet = True
        functionLibrarySheet.protection.enable()
        functionLibrarySheet.protection.password = '123'
        functionLibrarySheet.cell(1, 1).value = 'function_library'
        functionLibrarySheet.cell(1, 2).value = 'function_name'
        functionLibrarySheet.cell(1, 3).value = 'function_parameters'
        functionLibrarySheet.cell(1, 5).value = 'function library references'

    # * If function definition map is not empty, write new data to function definitions sheet and reapply defined names
    if functionDefinitionMap:
        
        # * Clear function definitions sheet of old data
        for row in functionLibrarySheet['A2:E1000']:
            for cell in row:
                cell.value = None

        # * Write new data to function definitions sheet
        row = 2
        for i, (functionLibrary, functionNames) in enumerate(functionDefinitionMap.items()):
            functionLibrarySheet.cell(i+2, 5).value = functionLibrary
            
            count = 0
            for j, (functionName, data) in enumerate(functionNames.items()):
                functionLibrarySheet.cell(row, 1).value = functionLibrary
                functionLibrarySheet.cell(row, 2).value = functionName
                functionLibrarySheet.cell(row, 3).value = '\n'.join(data['function_parameters'])

                try:
                    del workbook.defined_names[functionName]
                except Exception as ex:
                    print(ex)
                finally:
                    new_range = openpyxl.workbook.defined_name.DefinedName(functionName, attr_text=f"'function definitions'!$C${row}")
                    workbook.defined_names.append(new_range)

                row += 1
                count += 1

            try:
                del workbook.defined_names[functionLibrary]
            except Exception as ex:
                print(ex)
            finally:
                new_range = openpyxl.workbook.defined_name.DefinedName(functionLibrary, attr_text=f"'function definitions'!$B${row-count}:$B${row}")
                workbook.defined_names.append(new_range)

    workbook.save(xlsxOutFile)

'''
Handles the writing of updated function definitions data to the function definitions database,
these changes are made through the GUI.

The function recieves a functionDefinitionMap object containing the updated information for function definitions,
functionDefinitionInFile string which specifies the file path for the function defintion database.

No return Value, writes updated function defintion data to the existing database
'''
def handleFunctionDefinitionDataUpdate(functionDefinitionMap, functionDefinitionInFile):

    # * Load excel file with openpyxl load_workbook
    workbook = load_workbook(filename=functionDefinitionInFile)
    sheet = workbook['function definitions']

    # * Clear function definitions sheet of old data
    for row in sheet['A2:E1000']:
        for cell in row:
            cell.value = None

    # * Write new data to function definitions sheet
    row = 2
    for i, (functionLibrary, functionNames) in enumerate(functionDefinitionMap.items()):
        # Write function libary references to column E
        sheet.cell(i+2, 5).value = functionLibrary
        
        # Write function definition data to column A, B, C
        for functionName, data in functionNames.items():
            sheet.cell(row, 1).value = functionLibrary
            sheet.cell(row, 2).value = functionName
            sheet.cell(row, 3).value = '\n'.join(data['function_parameters'])

            row += 1

    workbook.save(functionDefinitionInFile)

# ********************************************************* Helper functions ********************************************************#

# * Helper function which helps to parse the config excel file mapping sheet for any empty fields in the translations
# * Used by the handleMappingData function
def getEmptyFieldData(conversionMap):
    emptyFieldData = {
        'title': 'Empty fields',
        'warning': config_warnings.EMPTY_MAPPING_FIELDS_WARNINGS,
        'worksheet': 'mapping',
        'data': {}
    }

    # * Iterate through conversion mapping
    # * 
    for cleanedOldDescription, mapping in conversionMap.items():
        row_location = conversionMap[cleanedOldDescription]['configRowCount']
        
        # * Check if there are any empty fields in the teststep
        emptyFields = {}
        for tag, value in mapping.items():

            if tag == "isMatched":
                continue

            if str(value).startswith('empty_description_key'):
                emptyFields['empty_description_key'] = f"A{row_location}"
            elif not value:
                emptyFields['DD2.0 ' + tag] = f"{HEADER_COLUMN_MAP[tag]}{row_location}"
        
        # * If there are empty fields and the whole config row is not fully empty
        # * Assign empty field data values to config role location key
        if emptyFields and len(emptyFields) < 5: 
            emptyFieldData['data'][row_location] = emptyFields

    return emptyFieldData

# * Helper function which helps to generate the classic teststep to DD2.0 teststep mapping pair for use to display in the GUI data grid.
# * Used by the handleTranslateXmlData function
def getTeststepData(index, teststep, mapping, cleanedOldDescription, oldDescription, childParentMap):
    # Get all teststep information
    oldFunctionLibrary = teststep.find("function_library").text
    oldFunctionName = teststep.find("function_name").text
    oldFunctionParams = teststep.find(
        "function_parameters").iter("param")

    old_function_parameters = {
        param.get('name'): param.text.strip('\n ')
        for param in oldFunctionParams
    }

    # Create old data object
    oldTestStepData = {
        "cleanedDescription": cleanedOldDescription,
        "description": oldDescription,
        "function_library": oldFunctionLibrary,
        "function_name": oldFunctionName,
        "function_parameters": old_function_parameters,
    }

    # Create new data object
    new_function_parameters = {}
    for param in mapping['function_parameters']:
        #* If function name is already in function parameters skip the parameter name
        if param['name'] in new_function_parameters:
            continue
        
        # * Use regex to filter out all reference parameter values enclosed by '@@'
        # * Try to access old paramter values that matches references
        # * If matches, add values to xml data
        # * Else, ignore it
        # Function to process each regex match for r'@@[^,]*@@'
        def matchOldParams(match):
            value = old_function_parameters.get(match.group().strip('@'))
            if value: return value

        # Generate processed parameter values after regex filtering
        parameter_values = []
        for text in [text for text in param['text'].split(',') if text]:
            value = re.sub(r'@@[^,]*@@', matchOldParams, text)
            if value: parameter_values.append(value)

        new_function_parameters[param['name']] = ','.join(parameter_values)
            
    newTestStepData = {
        "description": mapping["description"],
        "function_library": mapping["function_library"],
        "function_name": mapping["function_name"],
        "function_parameters": new_function_parameters,
    }

    # Append to teststep xmlData if teststep id has not been matched yet
    teststepData = {
        "id": index + 1,
        "parentId": childParentMap[teststep].get("id"),
        "parentType": childParentMap[teststep].tag,
        "parentName": childParentMap[teststep].get("name"),
        "configRowCount": mapping["configRowCount"],
        "old": oldTestStepData,
        "new": newTestStepData,
    }
    
    return teststepData
