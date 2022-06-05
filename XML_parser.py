import xml.etree.ElementTree as ET
from utils import *
from openpyxl import load_workbook

def handleXlsx(xlsxFile):
    #Load excel file with openpyxl load_workbook
    workbook = load_workbook(filename=xlsxFile)
    sheet = workbook.active

    #Convert excel sheet into a list of dictionary with header:value pairs
    reader = ExcelDictReader(sheet)

    #Convert dictionary list into atpMap
    atpMap = {}
    for row in reader:

        funcParams = row['function_parameters'].split('\n')
        newfuncParamsPair = []

        for param in funcParams:
            param = param.split('=')
            paramNameText = {
                'name': param[0],
                'text': param[1]
            }
            newfuncParamsPair.append(paramNameText)

        atpMap[row['teststep.desc old']] = {
            'description': row['test_step.desc new'],
            'function_library': row['function_library'],
            'function_name': row['function_name'],
            'function_parameters': newfuncParamsPair,
        }

    return atpMap

    
def getTestStepData(xmlInFile, atpMap):
    tree = ET.parse(xmlInFile)
    root = tree.getroot()
    childParentMap = {c: p for p in root.iter() for c in p}
    allTestSteps = root.iter('teststep')
    
    # Initialize xmlData list
    xmlData = []
    
    # Initialize counter to create id for each test step
    idCounter = 0
    
    for teststep in allTestSteps:
        
        
        # Get teststep description attribute from teststep
        oldDesciption = teststep.get('desc')
        
        if oldDesciption in atpMap:
            idCounter += 1
            oldFunctionLibrary = teststep.find('function_library').text
            oldFunctionName = teststep.find('function_name').text
            oldFunctionParams = teststep.find('function_parameters').iter('param')
            
            oldTestStepData = {
                'id': idCounter,
                'description': oldDesciption,
                'function_library': oldFunctionLibrary,
                'function_name': oldFunctionName,
                'function_parameters': [{
                    'name': param.get('name'),
                    'text': param.text.strip('\n ')
                    } for param in oldFunctionParams]
            }
            
            newTestStepData = {
                'id': idCounter,
                'description': atpMap[oldDesciption]['description'],
                'function_library': atpMap[oldDesciption]['function_library'],
                'function_name': atpMap[oldDesciption]['function_name'],
                'function_parameters': [{
                    'name': param['name'].strip(),
                    'text': param['text'].strip('\n ')
                    } for param in atpMap[oldDesciption]['function_parameters']]
            }
            
            xmlData.append({
                'parentId': childParentMap[teststep].get('id'),
                'parentName': childParentMap[teststep].get('name'),
                'old': oldTestStepData,
                'new': newTestStepData,
            })
        
    return xmlData


def convertXML(xmlInFile, xmlOutFile, atpMap):
    tree = ET.parse(xmlInFile)
    root = tree.getroot()
    allTestSteps = root.iter('teststep')

    for teststep in allTestSteps:
       
        # Get teststep description attribute from teststep
        oldtestStepDescription = teststep.attrib['desc']

        # IF teststep description finds match in atpMap - convert to new version
        if oldtestStepDescription in atpMap:
            matchPairs = atpMap[oldtestStepDescription]

            #Change old teststep description to new description
            teststep.set('desc', matchPairs['test_step.desc'])

            #Change old function_library text to new function_library text
            oldFunctionLibrary = teststep.find('function_library')
            oldFunctionLibrary.text = matchPairs['function_library']

            #Change old function_name text to new function_name text
            oldFunctionName = teststep.find('function_name')
            oldFunctionName.text = matchPairs['function_name']

            #Delete old function parameters and replace with new ones from atpMap
            oldFunctionParams = teststep.find('function_parameters')

            #remove existing param elements from function_parameters
            for param in list(oldFunctionParams.iter('param')):
                oldFunctionParams.remove(param)
               
            #create new param elements and append to function_parameters
            for param in matchPairs['function_parameters']:
                newParam = ET.SubElement(oldFunctionParams, 'param')
                newParam.set('name', param['name'])
                newParam.text = param['text']


            #Format newly populated function_parameters with proper indents and next lines
            print(f'''
                {teststep.get('desc')}
                {oldFunctionLibrary.text}
                {oldFunctionName.text}
                {[f"{param.get('name')}={param.text}" for param in oldFunctionParams]}
            ''')
            
    tree.write(xmlOutFile)


def testHandleXlsx():
    xlsxFile = './testdata/mapping.xlsx'
    
    atpMap = handleXlsx(xlsxFile)
    for key, value in atpMap.items():
        print(key)
        for key, value in value.items():
            print(f'{key}: {value}')
            print()
            
        break

def testHandleConvertXML():
    xlsxFile = './testdata/mapping.xlsx'
    xmlFile = './testdata/input.xml'

    atpMap = handleXlsx(xlsxFile)
    convertXML(xmlFile, './testdata/output.xml', atpMap)
    
def testHandleGetTestStepData():
    xlsxFile = './testdata/mapping.xlsx'
    xmlFile = './testdata/input.xml'
    
    atpMap = handleXlsx(xlsxFile)
    
    for item in getTestStepData(xmlFile, atpMap):
        print(item)
        for key, value in item.items():
            print(f'{key}: {value}')
            print()
        break
    
    
if __name__ == '__main__':
    testHandleGetTestStepData()
